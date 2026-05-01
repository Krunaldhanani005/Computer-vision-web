"""
fr_database.py — Unified Database for Face Recognition
Database: fr_surveillance_db

Collections:
    known_persons      – Registered known faces + encodings + dashboard stats
    known_attendance   – Attendance records per day per known person
    unknown_persons    – Unknown face trackers and stats
    unknown_snapshots  – Snapshots for unknown faces
    blacklist_persons  – Registered blacklist faces + encodings + stats
    blacklist_alerts   – Individual alert events for blacklist detections
    zones              – Polygon zone storage
"""

import datetime
import pytz
from pymongo import MongoClient, DESCENDING, ASCENDING

IST = pytz.timezone('Asia/Kolkata')

def to_ist(dt_utc) -> datetime.datetime:
    if not isinstance(dt_utc, datetime.datetime):
        return dt_utc
    if dt_utc.tzinfo is None:
        dt_utc = dt_utc.replace(tzinfo=pytz.utc)
    return dt_utc.astimezone(IST)

try:
    _client = MongoClient("mongodb://localhost:27017/", serverSelectionTimeoutMS=3000)
    _db = _client["fr_surveillance_db"]

    known_persons     = _db["known_persons"]
    known_attendance  = _db["known_attendance"]
    unknown_persons   = _db["unknown_persons"]
    unknown_snapshots = _db["unknown_snapshots"]
    blacklist_persons = _db["blacklist_persons"]
    blacklist_alerts  = _db["blacklist_alerts"]
    zones_col         = _db["zones"]

    # Indexes
    known_persons.create_index("name", unique=True)
    known_persons.create_index([("last_seen", DESCENDING)])

    known_attendance.create_index([("name", ASCENDING), ("date", ASCENDING)], unique=True)
    known_attendance.create_index("date")

    unknown_persons.create_index("temp_id", unique=True)
    unknown_persons.create_index([("last_seen", DESCENDING)])
    unknown_persons.create_index("date")

    unknown_snapshots.create_index("temp_id")
    unknown_snapshots.create_index([("timestamp", DESCENDING)])

    blacklist_persons.create_index("name", unique=True)

    blacklist_alerts.create_index([("alert_time", DESCENDING)])
    blacklist_alerts.create_index("name")

    print("[fr_database] Connected to fr_surveillance_db ✓")
except Exception as e:
    print(f"[fr_database] MongoDB error: {e}")
    known_persons = known_attendance = unknown_persons = unknown_snapshots = blacklist_persons = blacklist_alerts = zones_col = None

# ══════════════════════════════════════════════════════════════════════════════
#  REGISTRATION & ENCODINGS
# ══════════════════════════════════════════════════════════════════════════════

_MAX_EMBEDDINGS_PER_PERSON = 10   # cap stored embeddings per person


def insert_face(name: str, person_type: str, encoding: list) -> bool:
    """
    Store ArcFace embedding for a person.

    Embeddings are kept in an array field ``encodings`` (max
    _MAX_EMBEDDINGS_PER_PERSON). Multiple training images therefore each
    contribute their own embedding, giving recognition a best-of-N match
    instead of just the last uploaded sample.

    Backward-compat: old documents with a single ``encoding`` field are
    migrated to the array format on first write.
    """
    try:
        now = datetime.datetime.utcnow()
        col = blacklist_persons if person_type == "blacklist" else known_persons

        existing = col.find_one({"name": name})
        if existing:
            current_encs = existing.get("encodings", [])
            # Migrate legacy single-field documents on the fly
            if not current_encs and existing.get("encoding"):
                current_encs = [existing["encoding"]]

            if len(current_encs) >= _MAX_EMBEDDINGS_PER_PERSON:
                print(f"[fr_database] insert_face: {name} already has "
                      f"{len(current_encs)} embeddings (cap={_MAX_EMBEDDINGS_PER_PERSON}), skipping")
                return True  # not an error — person is already well represented

            if not existing.get("encodings") and existing.get("encoding"):
                # First write on a legacy doc — migrate single→array and append new
                col.update_one(
                    {"name": name},
                    {"$set": {"encodings": current_encs + [encoding]},
                     "$unset": {"encoding": ""}}
                )
            else:
                col.update_one({"name": name}, {"$push": {"encodings": encoding}})
        else:
            col.insert_one({
                "name":        name,
                "person_type": person_type,
                "encodings":   [encoding],   # array from the start
                "created_at":  now,
                "status":      "active",
                "total_detections": 0,
            })
        return True
    except Exception as e:
        print(f"[fr_database] insert_face error: {e}")
        return False


def get_all_faces():
    """
    Return one dict per stored embedding so the in-memory cache gets an entry
    for every embedding, enabling best-of-N matching per person.

    Handles both new (``encodings`` array) and legacy (``encoding`` scalar) docs.
    """
    faces = []
    for col in (known_persons, blacklist_persons):
        if col is None:
            continue
        for doc in col.find({}, {"_id": 0, "name": 1, "person_type": 1,
                                 "encodings": 1, "encoding": 1}):
            name  = doc.get("name", "Unknown")
            ptype = doc.get("person_type", "known")
            encs  = doc.get("encodings", [])
            # Backward-compat: single encoding field on old documents
            if not encs and doc.get("encoding"):
                encs = [doc["encoding"]]
            for enc in encs:
                faces.append({"name": name, "person_type": ptype, "encoding": enc})
    return faces

def delete_all_faces() -> int:
    count = 0
    if known_persons is not None:
        count += known_persons.delete_many({}).deleted_count
    if blacklist_persons is not None:
        count += blacklist_persons.delete_many({}).deleted_count
    return count

# ══════════════════════════════════════════════════════════════════════════════
#  KNOWN PERSONS & ATTENDANCE
# ══════════════════════════════════════════════════════════════════════════════

def upsert_known(name: str, confidence: float, snapshot_path: str):
    if known_persons is None: return
    now = datetime.datetime.utcnow()
    date_str = to_ist(now).strftime("%Y-%m-%d")

    known_persons.update_one(
        {"name": name},
        {
            "$set": {
                "last_seen":       now,
                "latest_snapshot": snapshot_path,
                "confidence":      confidence,
            },
            "$inc": {"total_detections": 1},
            "$addToSet": {"attendance_days": date_str},
            "$setOnInsert": {
                "created_at":  now,
                "person_type": "known",
                "status":      "active",
            },
        },
        upsert=True,
    )

    if known_attendance is not None:
        att_update = {
            "$setOnInsert": {"first_seen_today": now, "person_id": name},
            "$set":         {"last_seen_today": now},
            "$inc":         {"detection_count": 1},
        }
        if snapshot_path:
            att_update["$set"]["snapshot"] = snapshot_path
            att_update["$push"] = {
                "snapshots": {"$each": [snapshot_path], "$slice": -5}
            }
        known_attendance.update_one(
            {"name": name, "date": date_str},
            att_update,
            upsert=True
        )

def get_known_dashboard(search: str = "", sort_by: str = "last_seen", page: int = 1, per_page: int = 24):
    if known_persons is None: return [], 0
    query = {}
    if search: query["name"] = {"$regex": search, "$options": "i"}

    sort_map = {
        "last_seen": [("last_seen", DESCENDING)],
        "name_asc": [("name", ASCENDING)],
        "name_desc": [("name", DESCENDING)],
        "detections": [("total_detections", DESCENDING)],
    }
    sort_key = sort_map.get(sort_by, [("last_seen", DESCENDING)])

    total = known_persons.count_documents(query)
    skip = (page - 1) * per_page
    docs = list(known_persons.find(query, {"encoding": 0, "_id": 0}).sort(sort_key).skip(skip).limit(per_page))

    today = to_ist(datetime.datetime.utcnow()).strftime("%Y-%m-%d")
    for d in docs:
        att_days = d.get("attendance_days", [])
        d["present_count"] = len(att_days)
        d["is_present_today"] = today in att_days
        if "last_seen" in d and hasattr(d["last_seen"], "strftime"):
            d["last_seen"] = to_ist(d["last_seen"]).strftime("%d-%m-%Y %H:%M:%S IST")
        if "created_at" in d and hasattr(d["created_at"], "strftime"):
            d["first_seen"] = to_ist(d["created_at"]).strftime("%d-%m-%Y %H:%M:%S IST")

    return docs, total

def get_known_stats():
    if known_persons is None: return {}
    today = to_ist(datetime.datetime.utcnow()).strftime("%Y-%m-%d")
    return {
        "registered_faces": known_persons.count_documents({}),
        "recognized_today": known_persons.count_documents({"attendance_days": today}),
    }

# ── Attendance (day-wise known) ───────────────────────────────────────────────

def get_attendance_by_date(date_str: str, search: str = "", page: int = 1, per_page: int = 50):
    if known_attendance is None: return [], 0
    query = {"date": date_str}
    if search:
        query["name"] = {"$regex": search, "$options": "i"}
    total = known_attendance.count_documents(query)
    skip  = (page - 1) * per_page
    docs  = list(
        known_attendance.find(query, {"_id": 0, "encodings": 0})
        .sort("detection_count", DESCENDING)
        .skip(skip).limit(per_page)
    )
    for d in docs:
        if "first_seen_today" in d and hasattr(d["first_seen_today"], "strftime"):
            d["first_seen"] = to_ist(d["first_seen_today"]).strftime("%H:%M:%S")
        if "last_seen_today" in d and hasattr(d["last_seen_today"], "strftime"):
            d["last_seen"] = to_ist(d["last_seen_today"]).strftime("%H:%M:%S")
        snaps = d.get("snapshots", [])
        if snaps and not d.get("snapshot"):
            d["snapshot"] = snaps[-1]
    return docs, total

def get_attendance_dates() -> list:
    if known_attendance is None: return []
    return sorted(known_attendance.distinct("date"), reverse=True)

# ══════════════════════════════════════════════════════════════════════════════
#  UNKNOWN PERSONS & SNAPSHOTS
# ══════════════════════════════════════════════════════════════════════════════

def upsert_unknown(temp_id: str, snapshot_path: str):
    if unknown_persons is None: return
    now = datetime.datetime.utcnow()
    date_str = to_ist(now).strftime("%Y-%m-%d")

    unknown_persons.update_one(
        {"temp_id": temp_id},
        {
            "$setOnInsert": {"first_seen": now, "date": date_str},
            "$set": {"last_seen": now, "latest_snapshot": snapshot_path},
            "$inc": {"detection_count": 1}
        },
        upsert=True
    )
    
    if unknown_snapshots is not None and snapshot_path:
        unknown_snapshots.insert_one({
            "temp_id": temp_id,
            "snapshot_path": snapshot_path,
            "timestamp": now
        })

def get_unknown_dashboard(search: str = "", sort_by: str = "last_seen", page: int = 1, per_page: int = 50):
    if unknown_persons is None: return [], 0
    query = {}
    if search: query["temp_id"] = {"$regex": search, "$options": "i"}

    sort_map = {
        "last_seen": [("last_seen", DESCENDING)],
        "first_seen": [("first_seen", DESCENDING)],
        "detections": [("detection_count", DESCENDING)],
    }
    sort_key = sort_map.get(sort_by, [("last_seen", DESCENDING)])

    total = unknown_persons.count_documents(query)
    skip = (page - 1) * per_page
    docs = list(unknown_persons.find(query, {"_id": 0}).sort(sort_key).skip(skip).limit(per_page))

    for d in docs:
        if "last_seen" in d and hasattr(d["last_seen"], "strftime"):
            ist_dt = to_ist(d["last_seen"])
            d["date"] = ist_dt.strftime("%d-%m-%Y")
            d["time"] = ist_dt.strftime("%H:%M:%S IST")

    return docs, total

def get_unknown_stats():
    if unknown_persons is None: return {}
    today = to_ist(datetime.datetime.utcnow()).strftime("%Y-%m-%d")
    return {
        "total_unknown": unknown_persons.count_documents({}),
        "unknown_today": unknown_persons.count_documents({"date": today}),
    }

def get_unknown_by_date(date_str: str, page: int = 1, per_page: int = 50):
    if unknown_persons is None: return [], 0
    query = {"date": date_str}
    total = unknown_persons.count_documents(query)
    skip  = (page - 1) * per_page
    docs  = list(
        unknown_persons.find(query, {"_id": 0})
        .sort("detection_count", DESCENDING)
        .skip(skip).limit(per_page)
    )
    for d in docs:
        for key in ("first_seen", "last_seen"):
            if key in d and hasattr(d[key], "strftime"):
                d[key + "_fmt"] = to_ist(d[key]).strftime("%H:%M:%S")
    return docs, total

def get_unknown_dates() -> list:
    if unknown_persons is None: return []
    return sorted(unknown_persons.distinct("date"), reverse=True)

def delete_all_unknown_logs():
    if unknown_persons is None: return False
    try:
        unknown_persons.delete_many({})
        unknown_snapshots.delete_many({})
        return True
    except:
        return False

# ══════════════════════════════════════════════════════════════════════════════
#  BLACKLIST ALERTS & PERSONS
# ══════════════════════════════════════════════════════════════════════════════

def insert_alert(name: str, snapshot_path: str, camera_source: str = "webcam", zone_id: str = "default"):
    if blacklist_alerts is None: return
    now = datetime.datetime.utcnow()
    
    # Also update blacklist_persons stats
    if blacklist_persons is not None:
        blacklist_persons.update_one(
            {"name": name},
            {"$set": {"last_seen": now, "latest_snapshot": snapshot_path}, "$inc": {"total_detections": 1}}
        )

    blacklist_alerts.insert_one({
        "name": name,
        "snapshot": snapshot_path,
        "alert_time": now,
        "camera_source": camera_source,
        "zone_id": zone_id
    })

def get_blacklist_dashboard(search: str = "", date_filter: str = "", page: int = 1, per_page: int = 50):
    if blacklist_alerts is None: return [], 0
    query = {}
    if search: query["name"] = {"$regex": search, "$options": "i"}
    if date_filter:
        try:
            day_start = datetime.datetime.strptime(date_filter, "%Y-%m-%d")
            day_end = day_start + datetime.timedelta(days=1)
            query["alert_time"] = {"$gte": day_start, "$lt": day_end}
        except ValueError:
            pass

    total = blacklist_alerts.count_documents(query)
    skip = (page - 1) * per_page
    docs = list(blacklist_alerts.find(query, {"_id": 0}).sort("alert_time", DESCENDING).skip(skip).limit(per_page))

    for d in docs:
        if "alert_time" in d and hasattr(d["alert_time"], "strftime"):
            d["alert_time"] = to_ist(d["alert_time"]).strftime("%d-%m-%Y %H:%M:%S IST")

    return docs, total

def get_blacklist_stats():
    if blacklist_alerts is None: return {}
    today_start = datetime.datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    return {
        "total_alerts": blacklist_alerts.count_documents({}),
        "alerts_today": blacklist_alerts.count_documents({"alert_time": {"$gte": today_start}}),
        "unique_persons": len(blacklist_alerts.distinct("name")),
    }

def get_blacklist_dates() -> list:
    if blacklist_alerts is None: return []
    dates = set()
    for doc in blacklist_alerts.find({}, {"alert_time": 1, "_id": 0}):
        at = doc.get("alert_time")
        if at and hasattr(at, "strftime"):
            dates.add(to_ist(at).strftime("%Y-%m-%d"))
    return sorted(dates, reverse=True)


# ══════════════════════════════════════════════════════════════════════════════
#  DAILY SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def get_daily_summary(date_str: str) -> dict:
    summary = {
        "date":             date_str,
        "attendance_count": 0,
        "known_count":      0,
        "unknown_count":    0,
        "blacklist_count":  0,
    }
    try:
        dt     = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        dt_end = dt + datetime.timedelta(days=1)
    except ValueError:
        return summary

    if known_attendance is not None:
        summary["attendance_count"] = known_attendance.count_documents({"date": date_str})
    if known_persons is not None:
        summary["known_count"] = known_persons.count_documents({"attendance_days": date_str})
    if unknown_persons is not None:
        summary["unknown_count"] = unknown_persons.count_documents({"date": date_str})
    if blacklist_alerts is not None:
        summary["blacklist_count"] = blacklist_alerts.count_documents(
            {"alert_time": {"$gte": dt, "$lt": dt_end}}
        )
    return summary

def get_summary_dates() -> list:
    dates = set()
    if known_attendance is not None:
        dates.update(known_attendance.distinct("date"))
    if unknown_persons is not None:
        dates.update(unknown_persons.distinct("date"))
    if blacklist_alerts is not None:
        for doc in blacklist_alerts.find({}, {"alert_time": 1, "_id": 0}):
            at = doc.get("alert_time")
            if at and hasattr(at, "strftime"):
                dates.add(to_ist(at).strftime("%Y-%m-%d"))
    return sorted(dates, reverse=True)

# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def export_attendance_xlsx(date_str: str) -> bytes:
    """Generate attendance_YYYY-MM-DD.xlsx with Attendance / Unknown / Blacklist sheets."""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="363062")
        hdr_aln  = Alignment(horizontal="center")

        def _style_header(ws, headers):
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(1, col_idx)
                cell.font  = hdr_font
                cell.fill  = hdr_fill
                cell.alignment = hdr_aln

        def _autowidth(ws):
            for col in ws.columns:
                width = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(width + 4, 45)

        # ── Sheet 1: Attendance ──────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Attendance"
        _style_header(ws1, ["Name", "First Seen", "Last Seen", "Detection Count"])
        if known_attendance is not None:
            for doc in (known_attendance.find({"date": date_str}, {"_id": 0})
                        .sort("detection_count", DESCENDING)):
                first = (to_ist(doc["first_seen_today"]).strftime("%H:%M:%S")
                         if hasattr(doc.get("first_seen_today"), "strftime") else "")
                last  = (to_ist(doc["last_seen_today"]).strftime("%H:%M:%S")
                         if hasattr(doc.get("last_seen_today"), "strftime") else "")
                ws1.append([doc.get("name", ""), first, last, doc.get("detection_count", 0)])
        _autowidth(ws1)

        # ── Sheet 2: Unknown ─────────────────────────────────────────────────
        ws2 = wb.create_sheet("Unknown")
        _style_header(ws2, ["Unknown ID", "First Seen", "Last Seen", "Detection Count"])
        if unknown_persons is not None:
            for doc in (unknown_persons.find({"date": date_str}, {"_id": 0})
                        .sort("detection_count", DESCENDING)):
                first = (to_ist(doc["first_seen"]).strftime("%H:%M:%S")
                         if hasattr(doc.get("first_seen"), "strftime") else "")
                last  = (to_ist(doc["last_seen"]).strftime("%H:%M:%S")
                         if hasattr(doc.get("last_seen"), "strftime") else "")
                ws2.append([doc.get("temp_id", ""), first, last, doc.get("detection_count", 0)])
        _autowidth(ws2)

        # ── Sheet 3: Blacklist ───────────────────────────────────────────────
        ws3 = wb.create_sheet("Blacklist")
        _style_header(ws3, ["Name", "Alert Time", "Camera Source", "Zone"])
        if blacklist_alerts is not None:
            try:
                dt     = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                dt_end = dt + datetime.timedelta(days=1)
                for doc in (blacklist_alerts.find(
                        {"alert_time": {"$gte": dt, "$lt": dt_end}}, {"_id": 0})
                        .sort("alert_time", DESCENDING)):
                    t = (to_ist(doc["alert_time"]).strftime("%H:%M:%S")
                         if hasattr(doc.get("alert_time"), "strftime") else "")
                    ws3.append([doc.get("name", ""), t,
                                doc.get("camera_source", ""), doc.get("zone_id", "")])
            except ValueError:
                pass
        _autowidth(ws3)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        print(f"[export_xlsx] Error: {e}")
        return b""


def export_known_csv() -> str:
    if known_persons is None: return "name,last_seen,present_count,total_detections\n"
    rows = ["name,last_seen,present_count,total_detections"]
    for d in known_persons.find({}, {"_id": 0, "encoding": 0}):
        last = to_ist(d.get("last_seen", "")).strftime("%d-%m-%Y %H:%M:%S IST") if hasattr(d.get("last_seen"), "strftime") else ""
        rows.append(f"{d.get('name','')},{last},{len(d.get('attendance_days', []))},{d.get('total_detections',0)}")
    return "\n".join(rows)

def export_blacklist_csv() -> str:
    if blacklist_alerts is None: return "name,alert_time,camera_source,zone_id\n"
    rows = ["name,alert_time,camera_source,zone_id"]
    for d in blacklist_alerts.find({}, {"_id": 0}):
        t = to_ist(d.get("alert_time", "")).strftime("%d-%m-%Y %H:%M:%S IST") if hasattr(d.get("alert_time"), "strftime") else ""
        rows.append(f"{d.get('name','')},{t},{d.get('camera_source','')},{d.get('zone_id','')}")
    return "\n".join(rows)

# ══════════════════════════════════════════════════════════════════════════════
#  ZONES
# ══════════════════════════════════════════════════════════════════════════════

def save_polygon_zone(points: list, zone_id: str = "default") -> bool:
    if zones_col is None: return False
    try:
        zones_col.delete_many({"zone_id": zone_id})
        zones_col.insert_one({"zone_id": zone_id, "points": points, "created_at": datetime.datetime.utcnow()})
        return True
    except:
        return False

def load_polygon_zone(zone_id: str = "default"):
    if zones_col is None: return None
    doc = zones_col.find_one({"zone_id": zone_id})
    return doc["points"] if doc and "points" in doc else None

def delete_polygon_zone(zone_id: str = "default"):
    if zones_col is None: return
    zones_col.delete_many({"zone_id": zone_id})

